import openpyxl as xl  # calling openpyxl package as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):  # iteration starts from 2 because we don't need the name of the columns
        cell = sheet.cell(row, 3)  # showing the data of column number 3
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)  # putting the corrected price in a new column
        corrected_price_cell.value = corrected_price

    # now to make a chart select range of values

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4,
                       max_col=4)  # we are selecting the cell from row 2 to 4 and only column 4
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
